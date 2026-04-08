#!/usr/bin/env python3
"""
Diagnostic script to examine Excel file structure and content.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'data_pipeline'))

from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
import pandas as pd

file_path = r"D:\Projects\gmr-mro\test_data_pipeline\Data\Data\2025\INCREMENTAL DATA-3\2025\HMV250001260825\Material_8Q-IAZ HMV25-000126-0825.xlsx"

print(f"Diagnosing: {file_path}")
print(f"File exists: {os.path.exists(file_path)}")
print(f"File size: {os.path.getsize(file_path)} bytes\n")

# Try to get sheet names
print("=" * 80)
print("STEP 1: Get sheet names")
print("=" * 80)
try:
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        workbook_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
        print("Workbook.xml content (first 1000 chars):")
        print(workbook_xml[:1000])
        print("\n")
except Exception as e:
    print(f"Error reading workbook.xml: {e}\n")

# Try pandas
print("=" * 80)
print("STEP 2: Try pandas ExcelFile")
print("=" * 80)
try:
    xls = pd.ExcelFile(file_path)
    print(f"Sheet names from pandas: {xls.sheet_names}\n")
except Exception as e:
    print(f"Error from pandas: {e}\n")

# Try openpyxl
print("=" * 80)
print("STEP 3: Try openpyxl load_workbook")
print("=" * 80)
try:
    wb = load_workbook(file_path, data_only=False, rich_text=False)
    print(f"Sheet names from openpyxl: {wb.sheetnames}")
    print(f"Active sheet: {wb.active.title}\n")
except Exception as e:
    print(f"Error from openpyxl: {e}\n")

# Try reading specific sheets
print("=" * 80)
print("STEP 4: Read PRICING sheet with pandas")
print("=" * 80)
try:
    df = pd.read_excel(file_path, sheet_name='PRICING')
    print(f"Shape: {df.shape}")
    print(f"Columns: {list(df.columns)}")
    print(f"First few rows:\n{df.head()}\n")
except Exception as e:
    print(f"Error reading PRICING sheet: {e}\n")

# Try reading PRICING via openpyxl
print("=" * 80)
print("STEP 5: Read PRICING sheet with openpyxl")
print("=" * 80)
try:
    wb = load_workbook(file_path, data_only=False, rich_text=False)
    ws = wb['PRICING']
    print(f"Sheet dimensions: {ws.dimensions}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Try to read some data
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        data.append(row)
        if i < 5:  # Print first 5 rows
            print(f"Row {i}: {row}")
    
    print(f"\nTotal rows read: {len(data)}")
    print(f"Non-empty rows: {sum(1 for row in data if any(v is not None for v in row))}\n")
except Exception as e:
    print(f"Error reading PRICING sheet with openpyxl: {e}\n")

# Try reading from XML directly
print("=" * 80)
print("STEP 6: Extract from XML")
print("=" * 80)
try:
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        # List all worksheets
        worksheet_files = [f for f in zip_ref.namelist() if f.startswith('xl/worksheets/')]
        print(f"Worksheet files found: {worksheet_files}\n")
        
        # Try to read the first worksheet
        if worksheet_files:
            sheet_xml_content = zip_ref.read(worksheet_files[0]).decode('utf-8')
            print(f"First 1500 chars of worksheet XML:")
            print(sheet_xml_content[:1500])
except Exception as e:
    print(f"Error reading worksheets: {e}\n")

print("\n" + "=" * 80)
print("DIAGNOSIS COMPLETE")
print("=" * 80)
