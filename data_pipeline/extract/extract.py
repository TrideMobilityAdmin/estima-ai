import os
import pandas as pd
import yaml
from logs.Log_config import setup_logging
import re
from fuzzywuzzy import process
from difflib import SequenceMatcher
from pymongo import MongoClient
import numpy as np
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io
import zipfile
# Initialize logger
logger = setup_logging()
#'/home/CNLT7197/estima-ai/data_pipeline/config.yaml'
# Load config.yaml
def load_config(config_path=None):
    """Load configuration from YAML file."""
    if config_path is None:
        # Try to find config relative to the current file
        current_dir = os.path.dirname(os.path.abspath(__file__))
        config_path = os.path.join(current_dir, '..', 'config.yaml')
        
        # If not found, try alternate location
        if not os.path.exists(config_path):
            config_path = '/home/CNLT7197/estima-ai/data_pipeline/config.yaml'
            
    with open(config_path, "r") as file:
        return yaml.safe_load(file)

# Step 1: Establish Database Connection
def connect_to_database(db_uri: str, db_name: str):
    """
    Connect to MongoDB synchronously using pymongo and return the database object.
    """
    try:
        client = MongoClient(db_uri)
        db = client[db_name]  # Get database reference
        client.admin.command("ping")  # Check connection
        print(f"✅ Connected to database: {db_name}")
        return db
    except Exception as e:
        print(f"❌ Error connecting to MongoDB: {e}")
        logger.error("Error connecting to MongoDB", exc_info=True)
        return None


def convert_package_number(package_number):
    if not isinstance(package_number, str):
        return package_number
        
    match = re.search(r"HMV(\d{2})(\d{6})(\d{4})", package_number)
    if match:
        part1 = match.group(1)  # The first two digits after HMV
        part2 = match.group(2)  # The next six digits
        part3 = match.group(3)  # The last four digits
       
        return f"HMV{part1}/{part2}/{part3}"
    
    return package_number 

def clean_fly_hours(value):
    if isinstance(value, str):
        if ":" in value:  # If it's a time format (hh:mm), convert to float (optional)
            parts = value.split(":")
            value = float(parts[0]) + float(parts[1]) / 60  # Convert to decimal hours
        else:
            value = pd.to_numeric(value, errors='coerce')  # Convert to float if possible
    return value

def list_available_sheets(file_path):
    """
    List all available sheet names in an Excel file.
    Useful for debugging missing sheets.
    Tries multiple methods to handle corrupted files.
    """
    try:
        # Try pandas first
        excel_file = pd.ExcelFile(file_path)
        return excel_file.sheet_names
    except Exception as e:
        logger.debug(f"pandas.ExcelFile failed for {file_path}: {e}")
        try:
            # If pandas fails, try openpyxl directly
            wb = load_workbook(file_path, rich_text=False, data_only=False)
            sheets = wb.sheetnames
            logger.info(f"Successfully retrieved sheet names for {file_path} using openpyxl: {sheets}")
            return sheets
        except Exception as inner_e:
            logger.debug(f"openpyxl failed for {file_path}: {inner_e}")
            try:
                # Last resort: try with even more lenient settings
                wb = load_workbook(file_path, rich_text=False, data_only=False, keep_vba=False)
                sheets = wb.sheetnames
                logger.info(f"Successfully retrieved sheet names for {file_path} using openpyxl (lenient): {sheets}")
                return sheets
            except Exception as final_e:
                logger.error(f"Could not list sheets in {file_path}: {final_e}")
                return []

def read_excel_safely(file_path, sheet_name, engine="openpyxl"):
    """
    Read Excel file with fallback handling for corrupted cell references and XML.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to read
        engine (str): Engine to use (openpyxl, pyxlsb, xlrd)
    
    Returns:
        pd.DataFrame: DataFrame from the sheet, or empty DataFrame if failed
    """
    try:
        # First try standard pandas read_excel
        engine_kwargs = {'data_only': True} if engine == 'openpyxl' else {}
        excel_file = pd.ExcelFile(file_path, engine=engine, engine_kwargs=engine_kwargs)
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        logger.info(f"Successfully read {file_path}, sheet {sheet_name} using pandas. Shape: {df.shape}")
        return df
    except Exception as e:
        error_msg = str(e)
        logger.warning(f"Standard pandas read failed for {file_path}, sheet {sheet_name}: {error_msg}")
        
        # If it fails with cell reference or XML issues, try direct XML extraction
        if ("does not match pattern" in error_msg or 
            "invalid XML" in error_msg or 
            "cannot read worksheets" in error_msg.lower()) and engine == "openpyxl":
            
            logger.info(f"Attempting direct XML extraction for {file_path}...")
            df = read_excel_from_xml(file_path, sheet_name)
            
            if not df.empty:
                logger.info(f"Successfully extracted {file_path}, sheet {sheet_name} from XML. Shape: {df.shape}")
                return df
            else:
                logger.error(f"XML extraction returned empty DataFrame for {file_path}, sheet {sheet_name}")
                return pd.DataFrame()
        else:
            logger.error(f"Failed to read {file_path}, sheet {sheet_name}: {error_msg}")
            return pd.DataFrame()


def read_excel_from_xml(file_path, sheet_name):
    """
    Read Excel file directly from XML by extracting it from the .xlsx ZIP archive.
    This bypasses openpyxl's validation entirely and works with corrupted XML.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str): Name of the sheet to read
    
    Returns:
        pd.DataFrame: DataFrame from the sheet
    """
    try:
        import xml.etree.ElementTree as ET
        
        logger.info(f"Attempting to extract {sheet_name} from XML in {file_path}")
        
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # First, get the relationships to map sheet names to file names
            workbook_rels = {}
            try:
                rels_xml = zip_ref.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                rels_root = ET.fromstring(rels_xml)
                for rel in rels_root:
                    rel_id = rel.get('Id')
                    target = rel.get('Target')
                    workbook_rels[rel_id] = target
            except Exception as e:
                logger.debug(f"Could not read relationships: {e}")
            
            # Now read workbook.xml to find the sheet mapping
            workbook_xml = zip_ref.read('xl/workbook.xml').decode('utf-8')
            workbook_root = ET.fromstring(workbook_xml)
            
            # Find sheets
            ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheet_file = None
            for sheet in workbook_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
                sheet_name_attr = sheet.get('name')
                rid = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                
                logger.debug(f"Found sheet: {sheet_name_attr}, rid: {rid}")
                
                if sheet_name_attr and sheet_name_attr.lower() == sheet_name.lower():
                    # Map the relationship ID to the actual file
                    sheet_file = workbook_rels.get(rid)
                    logger.info(f"Matched sheet {sheet_name} to file {sheet_file}")
                    break
            
            if not sheet_file:
                logger.error(f"Could not find sheet {sheet_name} in workbook")
                return pd.DataFrame()
            
            # Read the actual sheet XML
            sheet_path = f'xl/{sheet_file}'
            try:
                sheet_xml_content = zip_ref.read(sheet_path).decode('utf-8')
            except KeyError:
                logger.error(f"Sheet file {sheet_path} not found in archive")
                return pd.DataFrame()
            
            # Read shared strings (must be inside the with block)
            shared_strings = {}
            try:
                strings_xml = zip_ref.read('xl/sharedStrings.xml').decode('utf-8')
                strings_root = ET.fromstring(strings_xml)
                for i, si in enumerate(strings_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si')):
                    text_elements = si.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    text = ''.join([t.text or '' for t in text_elements])
                    shared_strings[i] = text
                logger.info(f"Loaded {len(shared_strings)} shared strings")
            except Exception as e:
                logger.debug(f"No shared strings found: {e}")
        
        # Parse the sheet XML (outside the context manager is OK)
        try:
            sheet_root = ET.fromstring(sheet_xml_content)
        except ET.ParseError as e:
            logger.error(f"Failed to parse XML: {e}")
            # Try to fix common XML issues
            logger.info("Attempting to fix malformed XML...")
            sheet_xml_content = sheet_xml_content.replace('&', '&amp;')
            sheet_xml_content = sheet_xml_content.replace('&amp;amp;', '&amp;')
            try:
                sheet_root = ET.fromstring(sheet_xml_content)
            except ET.ParseError as e2:
                logger.error(f"Still unable to parse XML after fixes: {e2}")
                return pd.DataFrame()
        
        # Extract cell values from the sheet
        # Extract all rows
        data = []
        max_col = 0
        
        for row in sheet_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
            row_data = []
            for cell in row.findall('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
                cell_type = cell.get('t', 'n')  # Default to number
                
                # Get cell value
                value_elem = cell.find('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                value = None
                
                if value_elem is not None and value_elem.text:
                    if cell_type == 's':  # String reference
                        try:
                            str_index = int(value_elem.text)
                            value = shared_strings.get(str_index, value_elem.text)
                        except ValueError:
                            value = value_elem.text
                    else:
                        value = value_elem.text
                
                row_data.append(value)
            
            if row_data:
                max_col = max(max_col, len(row_data))
                if any(v is not None for v in row_data):
                    data.append(row_data)
        
        logger.info(f"Extracted {len(data)} rows from sheet XML")
        
        if data:
            # Normalize row lengths
            for row in data:
                while len(row) < max_col:
                    row.append(None)
            
            headers = data[0]
            df = pd.DataFrame(data[1:], columns=headers)
            logger.info(f"Created DataFrame with shape {df.shape}")
            return df
        else:
            logger.warning(f"No data extracted from XML")
            return pd.DataFrame()
            
    except Exception as e:
        logger.error(f"Error reading from XML: {e}", exc_info=True)
        return pd.DataFrame()

def get_best_match(target, candidates):
    """
    Find the best matching string from a list of candidates using SequenceMatcher.
    
    Args:
        target (str): The expected column name.
        candidates (list): List of candidate strings.
    
    Returns:
        tuple: (best match string, similarity score) or (None, 0) if no match found.
    """
    best_match = None
    best_score = 0

    for candidate in candidates:
        score = SequenceMatcher(None, target.lower(), candidate.lower()).ratio() * 100  # Convert to percentage
        if score > best_score:
            best_match = candidate
            best_score = score

    return (best_match, best_score) if best_score > 70 else (None, 0)  # Threshold set at 70%

def detect_header_row(df, expected_columns, max_rows_to_check=6):
    """
    Dynamically detect which row contains the header by checking for matches with expected columns.
    
    Args:
        df (pd.DataFrame): DataFrame with potential header rows.
        expected_columns (list): List of expected column names.
        max_rows_to_check (int): Maximum number of rows to check for headers.
    
    Returns:
        int or None: Best header row index or None if not found.
    """
    best_match_score = 0
    best_row_index = None

    # Check each of the first few rows
    for i in range(min(max_rows_to_check, len(df))):
        row_values = df.iloc[i].astype(str).fillna('').tolist()  # Convert row to list of strings

        # Skip rows with too many missing values
        if sum(pd.isna(df.iloc[i])) > len(row_values) / 2:
            continue

        # Calculate match score for this row
        row_score = 0
        for expected_col in expected_columns:
            best_match = get_best_match(expected_col, row_values)  # Find best match
            row_score += best_match[1]  # Add similarity score

        # If this row has better matches than previous best, update best row
        if row_score > best_match_score:
            best_match_score = row_score
            best_row_index = i

    return best_row_index

def predict_column_mappings(df, expected_mappings):
    """
    Dynamically predict column mappings using fuzzy matching.
    
    Args:
        df: DataFrame with the original column names
        expected_mappings: Dictionary of expected column names to their standardized names
    
    Returns:
        Dictionary mapping actual column names to standardized names
    """
    # Get all column names from the DataFrame
    actual_columns = list(df.columns)
    
    # Create a mapping for each expected column
    dynamic_mappings = {}
    
    for expected_col, standardized_name in expected_mappings.items():
        # Find the best match among actual columns
        best_match = process.extractOne(expected_col, actual_columns)
        
        if best_match and best_match[1] > 70:  # Match score above 70%
            dynamic_mappings[best_match[0]] = standardized_name
            logger.info(f"Mapped '{best_match[0]}' to '{standardized_name}' with score {best_match[1]}")
        else:
            logger.warning(f"Could not find a good match for '{expected_col}'")
    
    return dynamic_mappings

# Step 2: File Processing and Data Extraction
def get_processed_files(files_to_process, error_message=""):
    """
    Extract and combine data from Excel files into dataframes.
    Returns processed dataframes and a set of newly processed files.
    """
    try:
        config = load_config()  # Load once to avoid multiple reads
        newly_processed_files = set()

        def read_and_process_file(file_path, sheet_name, folder_name):
            """Read and process an individual Excel file."""
            try:
                file_extension = os.path.splitext(file_path)[1]
                engine = "openpyxl" if file_extension in [".xlsx", ".xlsm"] else "pyxlsb" if file_extension == ".xlsb" else "xlrd"

                # Read Excel file with safe handling for corrupted cell references
                df = read_excel_safely(file_path, sheet_name, engine)

                #print(f"df.shape: {df.shape}")
                #print(f"df.columns: {df.columns}")




                if sheet_name.lower() in ["pricing", "sheet1", "price", "page"]:

                    sub_task_parts_columns = config["sub_task_parts_columns"]
                    sub_task_parts_column_mappings = config["sub_task_parts_column_mappings"]

                    # Alternative mappings
                    alternative_mappings = {
                        "Issued Part#": "issued_part_number",
                        "Package#": "package_number",
                        "Task#": "task_number",
                        "SOI_TRANNO": "soi_transaction"
                    }

                    # Merge mappings
                    combined_mappings = {**sub_task_parts_column_mappings, **alternative_mappings}

                    # Detect the header row
                    header_row_index = detect_header_row(df, combined_mappings.keys())
                    print("df.shape:", df.shape, "of file:", file_path)

                    if header_row_index is not None:
                        df.columns = df.iloc[header_row_index].astype(str).values
                        print("df.columns:", df.columns, "of file:", file_path)

                        logger.info(f"Detected header row at index {header_row_index} for {file_path}, columns: {df.columns}")

                        # Extract data rows (everything after the header)
                        df = df.iloc[header_row_index + 1:].reset_index(drop=True)
                    else:
                        # If no header row is detected, use the first row as header and log a warning
                        
                        logger.warning(f"Could not detect header row in {file_path}")
                        print(f"Could not detect header row in {file_path}")




                    # Rename columns using mappings
                    df.rename(columns={k: v for k, v in combined_mappings.items() if k in df.columns}, inplace=True)

                    expected_output_columns = list(sub_task_parts_column_mappings.values())
                    mapped_columns = set(df.columns)
                    truly_missing = set(expected_output_columns) - mapped_columns
                    if len(truly_missing) > 4:
                        return pd.DataFrame()

                    if truly_missing:
                        warning_msg = f"⚠️ Warning: Missing columns in {file_path} that couldn't be mapped: {truly_missing}"
                        print(warning_msg)
                        logger.warning(warning_msg)

                    # Remove duplicate columns
                    df = df.loc[:, ~df.columns.duplicated()].copy()
                    # Ensure all expected columns exist
                    for col in expected_output_columns:
                        if col not in df.columns:
                            df[col] = None
                    # Convert to string
                    string_cols = [
                        'registration_number', 'package_number', 'task_number',
                        'task_description', 'issued_part_number', 'part_description',
                        'issued_unit_of_measurement', 'stock_status', 'base_currency',
                        'soi_transaction'
                    ]

                    # Apply conversions
                    for col in string_cols:
                        if col in df.columns:
                            # Replace NaN with None
                            df[col] = df[col].where(pd.notnull(df[col]), None)
                            
                            # Convert non-None values to string (but keep None as None)
                            df[col] = df[col].apply(lambda x: str(x) if x is not None else None)

                        
                    numeric_convert = ['base_price_usd', 'freight_cost', 'admin_charges', 'total_billable_price', 'billable_value_usd', 'used_quantity']

                    for col in numeric_convert:
                        df[col] = pd.to_numeric(df[col], errors='coerce')

                    # Reorder columns to match expected order
                    df = df[expected_output_columns]
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)
                    print("Processed sub_task_parts file:", file_path)
                    
                elif sheet_name == "HMV" or sheet_name[:2] == "20":
                    df.columns = df.columns.astype(str).str.strip()  # Strip spaces from column names
                    df = df.loc[:, ~df.columns.duplicated()].copy()  # Remove duplicate columns
                    df = df.reset_index(drop=True)  # Reset index

                    aircraft_details_column_mappings = config["aircraft_details_column_mappings"]
                    aircraft_details_columns = [col.strip() for col in config["aircraft_details_columns"]]

                    # Identify missing and extra columns
                    missing_cols = set(aircraft_details_columns) - set(df.columns)
                    extra_cols = set(df.columns) - set(aircraft_details_columns)

                    if missing_cols:
                        print(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                        logger.warning(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                    
                    if extra_cols:
                        print(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")
                        logger.warning(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")

                    df["year"] = int(folder_name)  # Add the 'year' column

                    # Rename columns if they exist in the dataframe
                    df.rename(columns={k: v for k, v in aircraft_details_column_mappings.items() if k in df.columns}, inplace=True)
                                        
                    # Add missing expected columns with None values
                    expected_columns = list(aircraft_details_column_mappings.values())
                    missing_columns = [col for col in expected_columns if col not in df.columns]
                    for col in missing_columns:
                        df[col] = None
                    df['flight_hours'] = df['flight_hours'].apply(clean_fly_hours)

                    # Handling aircraft age columns
                    for col in ["aircraft_age", "aircraft_age_2024"]:
                        if col in df.columns:
                            # Try to extract numeric values from strings, handle errors safely
                            df[col] = (
                                df[col]
                                .astype(str)  # Convert all values to string
                                .str.strip()  # Remove extra spaces
                                .str.extract(r'(\d+\.?\d*)')[0]  # Extract numeric part
                                .astype(float, errors='ignore')  # Convert to float with error handling
                            )
                            # Additional safety conversion with proper error handling
                            df[col] = pd.to_numeric(df[col], errors='coerce')


                    # Handle datetime columns
                    datetime_cols = ["issue_date", 'start_date', 'end_date']

                    for col in datetime_cols:
                        if col in df.columns:
                            # Convert to datetime with error handling
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                            # Only localize non-NaT values to avoid TypeError
                            mask = ~df[col].isna()
                            if mask.any():
                                df.loc[mask, col] = df.loc[mask, col].dt.tz_localize('UTC')
                            # Fill NaT values with a default timestamp (already timezone aware)
                            df[col] = df[col].fillna(pd.Timestamp('0001-01-01T00:00:00.000', tz='UTC'))

                    # Define string columns
                    string_cols = [
                        'customer_order_number', 'customer_name', 'package_number',
                        'aircraft_registration_number', 'aircraft_model',
                        'release_tracking_number', 'package_details', 'check_category',
                        'train_data', 'test_data', 'type_of_check', 'last_maintenance_date'
                    ]

                    # Convert string columns, handling None values

                    for col in string_cols:
                        if col in df.columns:
                            # Replace NaN with None
                            df[col] = df[col].where(pd.notnull(df[col]), None)
                            
                            # Convert non-None values to string (but keep None as None)
                            df[col] = df[col].apply(lambda x: str(x) if x is not None else None)


                    # Define numeric columns
                    numeric_cols = ['year', 'flight_hours', 'flight_cycles', 'aircraft_age', 
                                'aircraft_age_2024', 'turnaround_time']

                    # Convert numeric columns with error handling
                    for col in numeric_cols:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')



                    # Ensure the dataframe has all expected columns before reordering
                    reorder_columns = [col for col in expected_columns if col in df.columns]
                    df = df.reindex(columns=reorder_columns)

                    # Add file path column
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)
                    
                elif sheet_name == 'mlttable':
                    df.columns = df.iloc[0].astype(str).str.strip()
                    df = df[1:].reset_index(drop=True)

                    df = df.loc[:, ~df.columns.duplicated()].copy()
                    task_parts_columns_mappings = config["task_parts_columns_mappings"]
                    task_parts_columns = config["task_parts_columns"]
                    
                    # Ensure correct columns
                    missing_cols = set(task_parts_columns) - set(df.columns)
                    extra_cols = set(df.columns) - set(task_parts_columns)
                    
                    if missing_cols:
                        print(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                        logger.warning(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                    if extra_cols:
                        print(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")
                        logger.warning(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")
                        
                    df["package"] = folder_name  # Add folder name as a column
                    df['package'] = df['package'].apply(convert_package_number)
                    
                    # First rename the columns that exist
                    df.rename(columns={k: v for k, v in task_parts_columns_mappings.items() if k in df.columns}, inplace=True)

                    # Now add any missing columns (using the final mapped column names)
                    expected_columns = list(task_parts_columns_mappings.values())
                    missing_columns = [col for col in expected_columns if col not in df.columns]
                    string_cols=['rt_mode_flag', 'delete_flag', 'task_number', 'requested_part_number',
                    'issued_part_number', 'issued_serial_number', 'issued_lot_number',
                    
                    'part_description', 'requested_stock_status', 'material_request_number',
                    'part_type', 'issued_stock_status', 'issue_number',  'package_number']
                    # Add missing columns with None values
                    
                    # Apply conversions
                    for col in string_cols:
                        if col in df.columns:
                            # Replace NaN with None
                            df[col] = df[col].where(pd.notnull(df[col]), None)
                            
                            # Convert non-None values to string (but keep None as None)
                            df[col] = df[col].apply(lambda x: str(x) if x is not None else None)

                        
                    numeric_convert = ['issued_quantity', 'used_quantity', 'unit_of_measurement',
                    'pending_return_quantity', 'returned_quantity_excess','material_cost',
                    'requested_quantity']

                    for col in numeric_convert:
                        df[col] = pd.to_numeric(df[col], errors='coerce')
                        
                    for col in missing_columns:
                        df[col] = None

                    # Finally, reorder columns to match expected order
                    df = df[expected_columns]
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)
                    
                elif sheet_name == 'mltaskmlsec1':
                    df.columns = df.iloc[0].astype(str).str.strip()
                    df = df[1:].reset_index(drop=True)
                    df = df.loc[:, ~df.columns.duplicated()].copy()
                    task_description_columns = config["task_description_columns"]
                    task_description_columns_mappings = config["task_description_columns_mappings"]
                    missing_cols = set(task_description_columns) - set(df.columns)
                    extra_cols = set(df.columns) - set(task_description_columns)
                    
                    if missing_cols:
                        print(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                        logger.warning(f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}")
                    if extra_cols:
                        print(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")
                        logger.warning(f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}")
                        
                    df["package"] = folder_name
                    df['package'] = df['package'].apply(convert_package_number)
                    df.rename(columns={k: v for k, v in task_description_columns_mappings.items() if k in df.columns}, inplace=True)

                    # Now add any missing columns (using the final mapped column names)
                    expected_columns = list(task_description_columns_mappings.values())
                    missing_columns = [col for col in expected_columns if col not in df.columns]
                    
                    # Handle datetime columns
                    datetime_cols = [  'actual_start_date','actual_end_date']
                    for col in datetime_cols:
                        if col in df.columns:
                            # Convert to datetime with error handling
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                            # Only localize non-NaT values to avoid TypeError
                            mask = ~df[col].isna()
                            if mask.any():
                                df.loc[mask, col] = df.loc[mask, col].dt.tz_localize('UTC')
                            # Fill NaT values with a default timestamp (already timezone aware)
                            df[col] = df[col].fillna(pd.Timestamp('0001-01-01T00:00:00.000', tz='UTC'))

                    # Define string columns
                    string_cols = [
                        '', 'rt_mode_flag', 'delete_flag', 'sequence_number', 'task_number',
                        'ata_number', 'description', 'status', 'work_center_number',
                        'sign_off_status', 'task_type', 'task_category', 'job_type',
                        'execution_phase', 'execution_category', 'mechanic_required',
                        'inspection_required', 'rii_required',  'tracking_number',
                        'long_description', 'hold_status', 'estimation_status', 'skill_number',
                        'wbs_code', 'package_number', 'source_task_discrepancy_number',
                        'source_tracking_number', 'zone_number', 'parent_task_number',
                        'root_task_number', 'previous_execution_comments', 'mechanic_sign_off',
                        'inspection_sign_off', 'rii_sign_off', 'additional_sign_off',
                        'previous_sign_off_comments', 'workscoping_comments', 'repair_agency',
                        'repair_classification', 'maintenance_manual_reference_number'
                    ]
                    

                    # Convert string columns, handling None values
                    for col in string_cols:
                        if col in df.columns:
                            # Replace NaN with None
                            df[col] = df[col].where(pd.notnull(df[col]), None)
                            
                            # Convert non-None values to string (but keep None as None)
                            df[col] = df[col].apply(lambda x: str(x) if x is not None else None)


                    # Define numeric columns
                    numeric_cols = ['estimated_man_hours','actual_man_hours']

                    # Convert numeric columns with error handling
                    for col in numeric_cols:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        else:
                            df[col] = None
                    
                    

                    # Add missing columns with None values
                    for col in missing_columns:
                        df[col] = None

                    # Finally, reorder columns to match expected order
                    df = df[expected_columns]
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)
                
                elif sheet_name == 'mldpmlsec1':
                    # Ensure first row is used as column names safely
                    df.columns = df.iloc[0].astype(str).str.strip()
                    df = df[1:].reset_index(drop=True)
                    df = df.loc[:, ~df.columns.duplicated()].copy()

                    sub_task_description_columns = config["sub_task_description_columns"]
                    sub_task_description_columns_mappings = config["sub_task_description_columns_mappings"]

                    # Check for missing and extra columns
                    missing_cols = set(sub_task_description_columns) - set(df.columns)
                    extra_cols = set(df.columns) - set(sub_task_description_columns)
                    
                    if missing_cols:
                        warning_msg = f"⚠️ Warning: Missing columns in {file_path}: {missing_cols}"
                        print(warning_msg)
                        logger.warning(warning_msg)
                    
                    if extra_cols:
                        warning_msg = f"⚠️ Warning: Extra columns in {file_path}: {extra_cols}"
                        print(warning_msg)
                        logger.warning(warning_msg)

                    df["package"] = folder_name
                    df["package"] = df["package"].apply(convert_package_number)

                    # Rename columns safely
                    df.rename(columns={k: v for k, v in sub_task_description_columns_mappings.items() if k in df.columns}, inplace=True)

                    # Add any missing columns (based on mapped names)
                    expected_columns = list(sub_task_description_columns_mappings.values())
                    missing_columns = [col for col in expected_columns if col not in df.columns]
                    
                    # Handle datetime columns
                    datetime_cols = ['planned_start_date','planned_end_date',  'actual_start_date','actual_end_date']
                    for col in datetime_cols:
                        if col in df.columns:
                            # Convert to datetime with error handling
                            df[col] = pd.to_datetime(df[col], errors='coerce')
                            # Only localize non-NaT values to avoid TypeError
                            mask = ~df[col].isna()
                            if mask.any():
                                df.loc[mask, col] = df.loc[mask, col].dt.tz_localize('UTC')
                            # Fill NaT values with a default timestamp (already timezone aware)
                            df[col] = df[col].fillna(pd.Timestamp('0001-01-01T00:00:00.000', tz='UTC'))
                            
                    string_cols = ['', 'rt_mode_flag', 'delete_flag', 'task_type', 'log_item_number',
                    'ata_number', 'task_description', 'corrective_action',
                    'discrepancy_number', 'action_taken', 'task_status',
                    'source_task_discrepancy_number',
                    'source_task_discrepancy_number_updated', 'source_tracking_number',
                    'work_center_number', 'sign_off_status', 'contract_classification',
                    'task_category', 'execution_phase', 'execution_category',
                    'part_required', 'corrosion_related', 'major_item', 'is_repeat',
                    'reported_by', 'reported_date', 'mechanical_required',
                    'mechanical_sign_off', 'inspection_required', 'inspection_sign_off',
                    'rii_required', 'sequence_number', 'tracking_number', 'hold_status',
                    'estimation_status', 'deferral_calendar', 'deferral_type',
                    'package_number', 'zone_number', 'skill_number', 'additional_sign_off',
                    'part_number', 'serial_number', 'part_description', 'position_code',
                    'previous_sign_off_comments', 'radio_communication',
                    'mechanical_skill_number', 'inspection_skill_number', 'rii_sign_off',
                    'rii_skill_number']
                    # Convert string columns, handling None values
                    for col in string_cols:
                        if col in df.columns:
                            # Replace NaN with None
                            df[col] = df[col].where(pd.notnull(df[col]), None)
                            
                            # Convert non-None values to string (but keep None as None)
                            df[col] = df[col].apply(lambda x: str(x) if x is not None else None)

                    # Define numeric columns
                    numeric_cols = ['estimated_man_hours','actual_man_hours']

                    # Convert numeric columns with error handling
                    for col in numeric_cols:
                        if col in df.columns:
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                        else:
                            df[col] = None
                    # Add missing columns with None values
                    for col in missing_columns:
                        df[col] = None

                    # Ensure required columns exist before proceeding
                    required_columns = ["log_item_number", "source_task_discrepancy_number"]
                    for col in required_columns:
                        if col not in df.columns:
                            print(f"⚠️ Warning: Required column '{col}' is missing in {file_path}. Skipping task_findings processing.")
                            logger.warning(f"⚠️ Warning: Required column '{col}' is missing in {file_path}. Skipping task_findings processing.")
                            return pd.DataFrame()  # Return empty DataFrame if required columns are missing

                    # Initialize the new column
                    df["source_task_discrepancy_number_updated"] = ""

                    # Create task_findings_dict
                    findings = df["log_item_number"].tolist()
                    tasks = df["source_task_discrepancy_number"].tolist()
                    task_findings_dict = dict(zip(findings, tasks))

                    # Resolve task references safely (avoid infinite loops)
                    max_iterations = 10  # Safety limit
                    for finding in findings:
                        iteration = 0
                        current = finding
                        
                        while iteration < max_iterations:
                            if current not in task_findings_dict or task_findings_dict[current] == current:
                                break  # Stop if there's no further reference or self-referencing
                            next_value = task_findings_dict[current]
                            if next_value == finding:  # Circular reference detected
                                break
                            current = next_value
                            iteration += 1
                        
                        # Update with the resolved reference
                        task_findings_dict[finding] = current

                    # Assign resolved values back to DataFrame
                    df["source_task_discrepancy_number_updated"] = df["log_item_number"].map(task_findings_dict)

                    # Ensure correct column order
                    df = df[expected_columns]
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)
                    
                else:
                    df.columns = df.iloc[0].astype(str).str.replace(".", "", regex=False)
                    df = df[1:].reset_index(drop=True)
                    df["package"] = folder_name  # Add folder name as a column
                    df["file_path"] = file_path
                    df = df.where(pd.notnull(df), None)

                return df
                
            except Exception as e:
                logger.error(f"Error processing file {file_path}, sheet {sheet_name}: {e}")
                print(f"Error processing file {file_path}, sheet {sheet_name}: {e}")
                return pd.DataFrame()
                # Collecting data
                
        def validate_filename_keywords( filename: str) -> Optional[str]:
            """
            Validate if filename contains required keywords.
            
            Args:
                filename (str): Name of the file to validate
                
            Returns:
                Optional[str]: Expected file type if valid, None otherwise
            """
            filename_lower = filename.lower()
            
            if "mltaskmlsec1" in filename_lower:
                return "mltaskmlsec1"
            elif "mldpmlsec1" in filename_lower:
                return "mldpmlsec1"
            elif "material" in filename_lower:
                return "material"
            elif "mlttable" in filename_lower:
                return "mlttable"
            elif "aircraft" in filename_lower:
                return "aircraft_details"
            else:
                print(f"⚠️ Skipping file {filename}: does not match any expected keywords.")
                return None
        
        def files_check(file_path, sheet_names, folder_name,error_message):
            """
            Check and process individual files.
            
            Args:
                file_path (str): Full path to the file
                sheet_names (list): List of expected sheet names
                folder_name (str): Name of the folder containing the file
                
            Returns:
                pd.DataFrame: Processed dataframe or empty dataframe
            """
            print(f"Processing file: {file_path}")
            try:
                filename = os.path.basename(file_path)  # Extract filename from path
                file_extension = os.path.splitext(filename)[1].lower()
                
                if not file_extension:
                    file_extension = f".{filename.split('.')[-1].lower()}"
                
                # Check if file extension is supported
                if file_extension in [".xlsx", ".xlsm", ".xlsb", ".xls"]:
                    extension = file_extension
                else:
                    error_message += (
                        f", Unsupported file extension {file_extension} for {filename}. "
                        "Supported extensions are: .xlsx, .xlsm, .xlsb, .xls"
                    )
                    raise Exception(
                        f"Unsupported file extension {file_extension} for {file_path}. "
                        "Supported extensions are: .xlsx, .xlsm, .xlsb, .xls"
                    )
                
                # Read Excel file and get available sheets
                # Use the safe sheet listing function that handles corrupted files
                available_sheets = list_available_sheets(file_path)
                if not available_sheets:
                    print(f"⚠️ Could not read sheets from {file_path}")
                    error_message += f"\n File {file_path} could not be read or has no sheets. So, Skipping it."
                    return pd.DataFrame()
                
                sheet_names_lower = [name.lower() for name in sheet_names]
                valid_sheets = [sheet for sheet in available_sheets 
                            if sheet.lower() in sheet_names_lower]
        
                sheet_name = valid_sheets[0] if valid_sheets else None
                if not sheet_name:
                    print(f"⚠️ Skipping file {file_path}: None of the required sheets found.")
                    print(f"   Available sheets: {available_sheets}")
                    print(f"   Expected sheets: {sheet_names}")
                    error_message += f"\n File {file_path} does not contain any of the required sheets: {sheet_names}. Available sheets: {available_sheets}. So, Skipping it."
                    return pd.DataFrame()  # Return empty DataFrame if no valid sheet found

                # Process the file (assuming read_and_process_file function exists)
                df = read_and_process_file(file_path, sheet_name, folder_name)

                if df is not None and not df.empty:
                    df.reset_index(drop=True, inplace=True)
                    print(f"✅ Processed file: {file_path}")
                    return df
                else:
                    print(f"⚠️ No valid data found in file: {file_path}")
                    error_message += f"\n File {file_path} does not contain required columns names in the sheet {sheet_name}. So, Skipping it."
                    return pd.DataFrame()  

            except Exception as e:
                logger.error(f"❌ Error processing file: {file_path}: {e}")
                print(f"❌ Error processing file: {file_path}: {e}")
                return pd.DataFrame()
        


        def process_files(files_to_process,error_message):
            """
            Process multiple files based on their keywords.
            
            Args:
                files_to_process (list): List of file paths
                
            Returns:
                tuple: Five dataframes containing processed data
            """
            aircraft_details = pd.DataFrame()
            task_description = pd.DataFrame()
            task_parts = pd.DataFrame()
            sub_task_description = pd.DataFrame()
            sub_task_parts = pd.DataFrame()
            
            for file_path in files_to_process:
                # Skip any file path that contains '2019' in its directory structure
                if "2019" in os.path.normpath(file_path).split(os.sep):
                    continue
                
                # Extract filename and folder name from the full path
                filename = os.path.basename(file_path)
                folder_name = os.path.basename(os.path.dirname(file_path))
                
                # Validate filename keywords
                file_keyword = validate_filename_keywords(filename)
                
                if file_keyword:
                    if file_keyword == "mltaskmlsec1":
                        sheet_names = ["mltaskmlsec1"]
                        df = files_check(file_path, sheet_names, folder_name,error_message)
                        task_description = pd.concat([task_description, df], ignore_index=True)
                        
                    elif file_keyword == "mldpmlsec1":
                        sheet_names = ["mldpmlsec1"]
                        df = files_check(file_path, sheet_names, folder_name,error_message)
                        sub_task_description = pd.concat([sub_task_description, df], ignore_index=True)
                        
                    elif file_keyword == "material":
                        sheet_names = ["pricing", "Pricing", "sheet1", "price", "page"]
                        df = files_check(file_path, sheet_names, folder_name,error_message)
                        sub_task_parts = pd.concat([sub_task_parts, df], ignore_index=True)
                        
                    elif file_keyword == "mlttable":
                        sheet_names = ["mlttable"]
                        df = files_check(file_path, sheet_names, folder_name,error_message)
                        task_parts = pd.concat([task_parts, df], ignore_index=True)
                    elif file_keyword == "aircraft_details":
                        sheet_names = ["HMV", "hMV", "2024", "2025",'2023', '2022', '2021', '2020']
                        df = files_check(file_path, sheet_names, folder_name,error_message)
                        aircraft_details = pd.concat([aircraft_details, df], ignore_index=True)
                else:
                    #print(f"⚠️ Skipping file {file_path}: does not match any expected keywords.")
                    error_message += f"\n File {filename} in the folder {folder_name} does not have  any expected keywords like mltaskmlsec1, mldpmlsec1, material, mlttable,aircraft in the file name. So, Skipping it."
                    
            #print(error_message)
            print("Processing completed for all files.")
            # Return the processed dataframes
            return aircraft_details, task_description, task_parts, sub_task_description, sub_task_parts,error_message
        
        aircraft_details, task_description, task_parts, sub_task_description, sub_task_parts,error_message=process_files(files_to_process,error_message)
        return aircraft_details, task_description, task_parts, sub_task_description, sub_task_parts, newly_processed_files, error_message
    except Exception as e:
        error_message = f"Error in get_processed_files: {e}"
        logger.error(f"Error in get_processed_files: {e}")
        print(f"Error fetching processed files: {e}")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), set(),error_message